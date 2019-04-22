from openpyxl import *

wb = load_workbook("Transaction.xlsx")
sheet = wb["Sheet1"]

for i in range(2, sheet.max_row + 1):
    cell = sheet.cell(i, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(i, 4)
    corrected_price_cell.value = corrected_price

wb.save("Transaction_updated.xlsx")
